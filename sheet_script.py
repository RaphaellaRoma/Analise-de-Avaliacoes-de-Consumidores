from openpyxl import load_workbook
from web_scraping import lista, declarar_listas
from main import trope

#incializando os excel
wb = load_workbook('analise_de_filmes.xlsx')
wb.active = wb["Tabela de Filmes"]
ws = wb.active

i = 0
n = 0

#nomes na primeira linha da tabela
names =[
    "Filme",
    "Data de Lançamento",
    "Gênero",
    "Diretor",
    "Bilheteria(Gross USD)",
    "Consensu dos Críticos",
    "Consensu da Audiência",
    "Reviews dos Críticos",
    "Reviews da Audiência"
]

#inserção dos dados dos scripts para o excel sheet
for n in range(len(declarar_listas())):
    while i < (len(lista[0])):
        if n == 4 or n == 9 or n == 10 or n == 11:
            ws.cell(row=i+1, column=n+1)
        else:
            ws.cell(row=i+1, column=n+1, value=lista[n][i])
        i += 1
    i = 0
    n += 1

#deletando colunas vazias
ws.delete_cols(10,3)
ws.delete_cols(5,1)

#inserção dos nomes para a primeira linha
i = 0
while i < len(names):
    ws.cell(row=1, column=i+1, value=names[i])
    i += 1

#criação da coluna 5
ws.insert_cols(4, 1)
#inserção do nome "Trope" para a primeira linha da coluna 5
ws.cell(row=1, column=4, value="Trope")

#inserção de dados dos tropes para a coluna 5
i = 0
while i < len(trope):
    ws.cell(row=i+2, column=4, value=(trope[i].replace("trope", "")))
    i += 1

#salvando a tabela no excel
wb.save('analise_de_filmes.xlsx')